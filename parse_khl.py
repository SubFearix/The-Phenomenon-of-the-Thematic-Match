#!/usr/bin/env python3
"""
Парсер результатов матчей ХК «Сибирь» в сезоне КХЛ 2024/25.

Источник данных: https://www.khl.ru/calendar/1288/00/29/

Формат выходного Excel-файла:
  - Дата
  - Соперник
  - Дома/В гостях
  - Результат (Победа / Поражение)
  - Забили
  - Пропустили

Использование:
    pip install -r requirements.txt
    python parse_khl.py
"""

import re
import sys

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

URL = "https://www.khl.ru/calendar/1288/00/29/"
TEAM_NAME = "Сибирь"
OUTPUT_FILE = "sibir_results_2024_25.xlsx"


def fetch_page(url: str) -> str:
    """Загрузить HTML-страницу календаря КХЛ."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    }
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    response.encoding = "utf-8"
    return response.text


def parse_matches(html: str) -> list[dict]:
    """
    Извлечь результаты матчей из HTML-страницы календаря КХЛ.

    Возвращает список словарей с ключами:
        date, opponent, location, result, goals_for, goals_against
    """
    soup = BeautifulSoup(html, "lxml")
    matches: list[dict] = []

    # ------------------------------------------------------------------
    # Стратегия 1: поиск элементов-карточек матчей
    # ------------------------------------------------------------------
    card_selectors = [
        "li[class*='game']",
        "div[class*='game-card']",
        "div[class*='match-card']",
        "div[class*='GameCard']",
        "div[class*='calendar-game']",
        "a[class*='game']",
        "a[href*='/game/']",
    ]

    cards = []
    for sel in card_selectors:
        cards = soup.select(sel)
        if cards:
            break

    if cards:
        current_date = ""
        for card in cards:
            match = _parse_card(card, current_date)
            if match:
                if match.get("date"):
                    current_date = match["date"]
                matches.append(match)
        if matches:
            return matches

    # ------------------------------------------------------------------
    # Стратегия 2: поиск таблицы с результатами
    # ------------------------------------------------------------------
    tables = soup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        for row in rows:
            cols = row.find_all(["td", "th"])
            match = _parse_table_row(cols)
            if match:
                matches.append(match)
    if matches:
        return matches

    # ------------------------------------------------------------------
    # Стратегия 3: текстовый разбор
    # ------------------------------------------------------------------
    matches = _parse_text(soup.get_text(separator="\n"))
    return matches


def _find_final_score(text: str) -> re.Match | None:
    """
    Найти итоговый счёт матча в тексте.

    В КХЛ ничьих не бывает: при ничьей в основное время матч
    продолжается в овертайме (ОТ) или серии буллитов (Б).
    Если в тексте несколько счётов, возвращается последний
    (итоговый) с неравным результатом. Если все счета ничейные,
    возвращается последний найденный.
    """
    score_re = re.compile(r"(\d+)\s*[:–\-]\s*(\d+)")
    all_matches = list(score_re.finditer(text))
    if not all_matches:
        return None

    # Предпочитаем последний счёт с неравным результатом (итоговый)
    for m in reversed(all_matches):
        if m.group(1) != m.group(2):
            return m

    # Все счета ничейные — вернуть последний
    return all_matches[-1]


def _parse_card(card, fallback_date: str) -> dict | None:
    """Попытаться извлечь данные из HTML-карточки матча."""

    text = card.get_text(separator=" ", strip=True)

    # Ищем счёт вида «3 : 2» или «3:2» или «3 - 2»
    score_match = _find_final_score(text)
    if not score_match:
        return None

    score_home = int(score_match.group(1))
    score_away = int(score_match.group(2))

    # Ищем дату
    date_str = ""
    date_el = card.find(class_=re.compile(r"date|time|day", re.I))
    if date_el:
        date_str = date_el.get_text(strip=True)
    if not date_str:
        date_match = re.search(r"\d{2}\.\d{2}\.\d{4}", text)
        if date_match:
            date_str = date_match.group(0)
    if not date_str:
        date_str = fallback_date

    # Ищем названия команд
    team_els = card.find_all(class_=re.compile(r"team|club", re.I))
    teams = [el.get_text(strip=True) for el in team_els if el.get_text(strip=True)]

    if len(teams) < 2:
        # Пытаемся разделить текст по счёту
        before = text[: score_match.start()].strip()
        after = text[score_match.end() :].strip()
        teams = [before, after]

    if len(teams) < 2:
        return None

    home_team = teams[0]
    away_team = teams[1]

    return _build_match_dict(date_str, home_team, away_team, score_home, score_away)


def _parse_table_row(cols) -> dict | None:
    """Попытаться извлечь данные из строки таблицы."""
    texts = [c.get_text(strip=True) for c in cols]
    if len(texts) < 3:
        return None

    # Ищем ячейку со счётом
    for i, t in enumerate(texts):
        score_match = _find_final_score(t)
        if score_match:
            score_home = int(score_match.group(1))
            score_away = int(score_match.group(2))
            home_team = texts[i - 1] if i > 0 else ""
            away_team = texts[i + 1] if i + 1 < len(texts) else ""
            date_str = texts[0] if i > 1 else ""
            if TEAM_NAME in home_team or TEAM_NAME in away_team:
                return _build_match_dict(
                    date_str, home_team, away_team, score_home, score_away
                )
    return None


def _parse_text(text: str) -> list[dict]:
    """Резервный текстовый парсер: ищем паттерны матчей в тексте."""
    matches: list[dict] = []
    lines = text.split("\n")

    current_date = ""
    date_pattern = re.compile(
        r"(\d{1,2}\s+"
        r"(?:января|февраля|марта|апреля|мая|июня|"
        r"июля|августа|сентября|октября|ноября|декабря)"
        r"(?:\s+\d{4})?)"
        r"|(\d{2}\.\d{2}\.\d{4})"
    )
    game_pattern = re.compile(
        r"(.+?)\s+(\d+)\s*[:–\-]\s*(\d+)"
        r"(?:\s*(?:ОТ|OT|Б|SO|БУЛ)\s+(\d+)\s*[:–\-]\s*(\d+))?"
        r"\s+(.+)"
    )

    for line in lines:
        line = line.strip()
        if not line:
            continue

        dm = date_pattern.search(line)
        if dm:
            current_date = dm.group(0).strip()

        gm = game_pattern.search(line)
        if gm:
            home_team = gm.group(1).strip()
            score_home = int(gm.group(2))
            score_away = int(gm.group(3))
            # Если есть итоговый счёт после ОТ/Б, используем его
            if gm.group(4) and gm.group(5):
                score_home = int(gm.group(4))
                score_away = int(gm.group(5))
            away_team = gm.group(6).strip()
            if TEAM_NAME in home_team or TEAM_NAME in away_team:
                match = _build_match_dict(
                    current_date, home_team, away_team, score_home, score_away
                )
                if match:
                    matches.append(match)

    return matches


def _build_match_dict(
    date_str: str,
    home_team: str,
    away_team: str,
    score_home: int,
    score_away: int,
) -> dict | None:
    """Сформировать словарь с результатом матча для «Сибири»."""
    if TEAM_NAME in home_team:
        location = "Домашний"
        opponent = away_team
        goals_for = score_home
        goals_against = score_away
    elif TEAM_NAME in away_team:
        location = "Гостевой"
        opponent = home_team
        goals_for = score_away
        goals_against = score_home
    else:
        return None

    if goals_for > goals_against:
        result = "Победа"
    else:
        result = "Поражение"

    return {
        "date": date_str,
        "opponent": opponent,
        "location": location,
        "result": result,
        "goals_for": goals_for,
        "goals_against": goals_against,
    }


def create_excel(matches: list[dict], filename: str) -> None:
    """Создать Excel-файл с результатами матчей."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты Сибири 24-25"

    # Стили заголовков
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Дата",
        "Соперник",
        "Дома/В гостях",
        "Результат",
        "Забили",
        "Пропустили",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Заливка для побед и поражений
    win_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    loss_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    for row_idx, match in enumerate(matches, 2):
        values = [
            match["date"],
            match["opponent"],
            match["location"],
            match["result"],
            match["goals_for"],
            match["goals_against"],
        ]
        fill = win_fill if match["result"] == "Победа" else loss_fill
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center
            cell.border = thin_border
            cell.fill = fill

    # Автоширина столбцов
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 4

    wb.save(filename)
    print(f"Файл сохранён: {filename}")


def main() -> None:
    print(f"Загрузка страницы: {URL}")
    try:
        html = fetch_page(URL)
    except requests.RequestException as exc:
        print(f"Ошибка при загрузке страницы: {exc}", file=sys.stderr)
        sys.exit(1)

    print("Разбор данных…")
    matches = parse_matches(html)

    if not matches:
        print(
            "Матчи не найдены. Возможно, структура сайта изменилась.\n"
            "Проверьте HTML-код страницы и при необходимости обновите парсер.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Найдено матчей: {len(matches)}")
    create_excel(matches, OUTPUT_FILE)
    print("Готово!")


if __name__ == "__main__":
    main()
